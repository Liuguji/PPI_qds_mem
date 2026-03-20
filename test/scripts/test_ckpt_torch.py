#!/usr/bin/env python3
"""
对比测试 flow_static_int_ckpt_torch 与 flow_static_int_ckpt_liubo 的数值一致性。

用法:
    python test/scripts/test_ckpt_torch.py
    python test/scripts/test_ckpt_torch.py --L 4 --qmax 300 --lmax 50
    python test/scripts/test_ckpt_torch.py --torch-only   # 只跑 torch 版（跳过对比）
"""

from __future__ import annotations

import os
import sys
import argparse
import time
import csv
from pathlib import Path
from datetime import datetime

os.environ.setdefault("JAX_ENABLE_X64", "true")
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")

REPO_ROOT = Path(__file__).resolve().parents[2]
CODE_DIR  = REPO_ROOT / "code"
sys.path.insert(0, str(CODE_DIR))

import unicodedata
import numpy as np
import models.models as models
from core.diag_routines.spinless_fermion import (
    flow_static_int_ckpt_liubo,
    flow_static_int_ckpt_torch,
)

W = 80  # 分隔线宽度（终端列数）


def _safe_token(v) -> str:
    """将参数值转换为适合文件名的 token。"""
    s = str(v)
    return ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in s)


def _write_per_l_result_file(result: dict, args) -> Path:
    """每个 L 生成一个结果文件：优先 xlsx，缺少 openpyxl 时回退 csv。"""
    out_dir = REPO_ROOT / "time_data_CPU_GPU"
    out_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    l_token = _safe_token(result.get("L", "unknown"))
    q_token = _safe_token(args.qmax)
    lmax_token = _safe_token(args.lmax)
    stem = f"ckpt_timing_L{l_token}_lmax{lmax_token}_qmax{q_token}_{ts}"

    headers = ["阶段", "CPU版本(JAX) 时间(s)", "GPU版本(Torch) 时间(s)"]
    stage_rows = [
        ["Part1 初始化", result.get("jax_part1_init_s"), result.get("torch_part1_init_s")],
        ["Part2 检查点配置", result.get("jax_part2_ckpt_s"), result.get("torch_part2_ckpt_s")],
        ["Part3 哈密顿量对角化", result.get("jax_forward_diag_s"), result.get("torch_forward_diag_s")],
        ["Part4 LIOM反向演化", result.get("jax_backward_liom_s"), result.get("torch_backward_liom_s")],
        ["总耗时", result.get("t_jax"), result.get("t_torch")],
    ]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        out_path = out_dir / f"{stem}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "ckpt_timing"

        ws.append(headers)
        for r in stage_rows:
            ws.append(r)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for col_idx, h in enumerate(headers, start=1):
            values = [str(h)] + [str(r[col_idx - 1]) for r in stage_rows]
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(v) for v in values) + 2, 40)

        wb.save(out_path)
        return out_path
    except Exception:
        out_path = out_dir / f"{stem}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(stage_rows)
        return out_path


def _dw(s: str) -> int:
    """字符串的终端显示宽度：中文等宽字符占 2 列，其余占 1 列。"""
    return sum(2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1 for c in s)


def _ljust_dw(s: str, width: int) -> str:
    """按终端显示宽度左对齐填充空格。"""
    return s + ' ' * max(width - _dw(s), 0)


def row(label: str, value: str, width: int = 24):
    print("  " + _ljust_dw(label, width) + " " + value)


def box_row(label: str, value: str, width: int = 18, border: str = "│"):
    content = "  " + _ljust_dw(label, width) + " " + value
    print(border + _ljust_dw(content, W - 2) + border)


def _print_result_block(label: str, elapsed: float, steps: int,
                        offdiag: float, cutoff: float, trace: float,
                        passed=None, lbits=None):
    print("┌" + "─" * (W - 2) + "┐")
    print("│" + _ljust_dw(f"  {label}", W - 2) + "│")
    print("├" + "─" * (W - 2) + "┤")
    box_row("耗时",            f"{elapsed:.2f} s")
    box_row("积分步数",        f"{steps}")
    box_row("off-diag 最大值", f"{offdiag:.3e}  (cutoff={cutoff:.1e})")
    box_row("LIOM2 迹",        f"{trace:.6f}  (期望 ≈ 1.0)")
    if lbits is not None:
        box_row("lbits", f"[{'  '.join(f'{v:.2f}' for v in lbits)}]")
    if passed is not None:
        print("├" + "─" * (W - 2) + "┤")
        box_row("判断", "✓ PASS" if passed else "✗ WARN")
    print("└" + "─" * (W - 2) + "┘")


def make_dl_list(lmax: float, qmax: int) -> np.ndarray:
    return np.logspace(np.log10(0.001), np.log10(lmax), qmax, endpoint=True, base=10)


def run_one(L: int, dim: int, qmax: int, lmax: float, cutoff: float,
            dis: float, dis_type: str, torch_only: bool) -> dict:
    n     = L ** dim
    J     = 1.0
    delta = 0.1
    x     = 0.0

    print("═" * (W + 10))
    print(f"  L={L}  n={n}  dim={dim}  dis={dis}  U={delta}  J={J}"
          f"  qmax={qmax}  lmax={lmax}  cutoff={cutoff:.1e}")
    print("═" * (W + 10))

    ham = models.hamiltonian("spinless fermion", dis_type, intr=True)
    ham.build(n, dim, dis, J, x, delta=delta)
    dl_list = make_dl_list(lmax, qmax)

    
    # ── torch 版 ──────────────────────────────────────────────────────────────
    t0 = time.perf_counter()
    res_torch = flow_static_int_ckpt_torch(
        n, ham, dl_list.copy(), qmax, cutoff, norm=False, Hflow=False,
    )
    t_torch = time.perf_counter() - t0

    H0_t      = res_torch["H0_diag"]
    liom2_t   = res_torch["LIOM2"]
    lbits_t   = res_torch["LIOM Interactions"]
    offdiag_t = float(np.max(np.abs(H0_t - np.diag(np.diag(H0_t)))))
    trace_t   = float(np.trace(liom2_t))

    _print_result_block("torch 版结果", t_torch, len(res_torch["dl_list"]),
                        offdiag_t, cutoff, trace_t, lbits=lbits_t)

    timing_torch = res_torch.get("_timing", {})
    torch_part1_init_s = float(timing_torch.get("part1_init_s", np.nan))
    torch_part2_ckpt_s = float(timing_torch.get("part2_ckpt_s", np.nan))
    torch_forward_diag_s = float(timing_torch.get("forward_diag_s", np.nan))
    torch_backward_liom_s = float(timing_torch.get("backward_liom_s", np.nan))
    row("torch Part1 初始化", f"{torch_part1_init_s:.3f} s")
    row("torch Part2 检查点", f"{torch_part2_ckpt_s:.3f} s")
    row("torch 对角化耗时", f"{torch_forward_diag_s:.3f} s")
    row("torch LIOM反向耗时", f"{torch_backward_liom_s:.3f} s")

    passed_torch = offdiag_t < cutoff * 10 and abs(trace_t - 1.0) < 0.5
    row("判断", f"{'✓ PASS' if passed_torch else '✗ WARN'}")

    if torch_only:
        return {"L": L, "n": n, "pass": passed_torch,
                "offdiag_torch": offdiag_t, "trace_torch": trace_t,
                "t_torch": t_torch,
            "torch_part1_init_s": torch_part1_init_s,
            "torch_part2_ckpt_s": torch_part2_ckpt_s,
                "torch_forward_diag_s": torch_forward_diag_s,
                "torch_backward_liom_s": torch_backward_liom_s}

    
    # ── JAX/liubo 版（对比基准）────────────────────────────────────────────────
    t0 = time.perf_counter()
    res_jax = flow_static_int_ckpt_liubo(
        n, ham, dl_list.copy(), qmax, cutoff, norm=False, Hflow=False,
    )
    t_jax = time.perf_counter() - t0

    H0_j      = res_jax["H0_diag"]
    liom2_j   = res_jax["LIOM2"]
    offdiag_j = float(np.max(np.abs(H0_j - np.diag(np.diag(H0_j)))))
    trace_j   = float(np.trace(liom2_j))

    _print_result_block("JAX (liubo) 版结果", t_jax, len(res_jax["dl_list"]),
                        offdiag_j, cutoff, trace_j)

    timing_jax = res_jax.get("_timing", {})
    jax_part1_init_s = float(timing_jax.get("part1_init_s", np.nan))
    jax_part2_ckpt_s = float(timing_jax.get("part2_ckpt_s", np.nan))
    jax_forward_diag_s = float(timing_jax.get("forward_diag_s", np.nan))
    jax_backward_liom_s = float(timing_jax.get("backward_liom_s", np.nan))
    row("JAX Part1 初始化", f"{jax_part1_init_s:.3f} s")
    row("JAX Part2 检查点", f"{jax_part2_ckpt_s:.3f} s")
    row("JAX 对角化耗时", f"{jax_forward_diag_s:.3f} s")
    row("JAX LIOM反向耗时", f"{jax_backward_liom_s:.3f} s")

    
    
    # ── 数值对比 ──────────────────────────────────────────────────────────────
    eig_err  = float(np.max(np.abs(np.sort(np.diag(H0_t)) - np.sort(np.diag(H0_j)))))
    liom_err = float(np.linalg.norm(liom2_t - liom2_j))

    numerically_close = eig_err < 5e-3 and liom_err < 5e-2
    row("数值一致性", f"{'✓ OK' if numerically_close else '✗ DIFF'}")

    return {
        "L": L, "n": n,
        "offdiag_torch": offdiag_t, "trace_torch": trace_t,
        "offdiag_jax":   offdiag_j, "trace_jax":   trace_j,
        "eig_err": eig_err, "liom_err": liom_err,
        "t_torch": t_torch, "t_jax": t_jax,
        "torch_part1_init_s": torch_part1_init_s,
        "torch_part2_ckpt_s": torch_part2_ckpt_s,
        "torch_forward_diag_s": torch_forward_diag_s,
        "torch_backward_liom_s": torch_backward_liom_s,
        "jax_part1_init_s": jax_part1_init_s,
        "jax_part2_ckpt_s": jax_part2_ckpt_s,
        "jax_forward_diag_s": jax_forward_diag_s,
        "jax_backward_liom_s": jax_backward_liom_s,
        "pass": passed_torch and numerically_close,
    }


def main():
    parser = argparse.ArgumentParser(
        description="对比 flow_static_int_ckpt_torch 与 flow_static_int_ckpt_liubo",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python test/scripts/test_ckpt_torch.py
  python test/scripts/test_ckpt_torch.py --L 4 6
  python test/scripts/test_ckpt_torch.py --torch-only --L 4
        """,
    )
    parser.add_argument("--L",          type=int,   nargs="+", default=[4],   help="线性系统大小（默认 4）")
    parser.add_argument("--dim",        type=int,   default=2,                help="空间维度（默认 2）")
    parser.add_argument("--qmax",       type=int,   default=1000,             help="最大流方程步数（默认 1000）")
    parser.add_argument("--lmax",       type=float, default=100.0,            help="最大流时间（默认 100）")
    parser.add_argument("--cutoff",     type=float, default=1e-3,             help="非对角截断值（默认 1e-3）")
    parser.add_argument("--dis",        type=float, default=5.0,              help="无序强度（默认 5.0）")
    parser.add_argument("--dis-type",   type=str,   default="random",         dest="dis_type")
    parser.add_argument("--torch-only", action="store_true",                  help="只运行 torch 版，跳过 JAX 对比")
    args = parser.parse_args()

    print("╔" + "═" * (W - 2) + "╗")
    print("║" + _ljust_dw("  测试目标: flow_static_int_ckpt_torch 与 flow_static_int_ckpt_liubo", W - 2) + "║")
    print("╠" + "═" * (W - 2) + "╣")
    box_row("L列表",    str(args.L),                                                          border="║")
    box_row("dim",       str(args.dim),                                                        border="║")
    box_row("qmax",      str(args.qmax),                                                       border="║")
    box_row("lmax",      str(args.lmax),                                                       border="║")
    box_row("cutoff",    f"{args.cutoff:.1e}",                                                 border="║")
    box_row("dis",       str(args.dis),                                                        border="║")
    box_row("模式",      "torch-only" if args.torch_only else "torch(GPU)vsJAX(CPU) 对比",   border="║")
    box_row("CKPT_STEP", os.environ.get("PYFLOW_CKPT_STEP", "auto, min(20, sqrt(qmax))"),     border="║")
    print("╚" + "═" * (W - 2) + "╝")

    all_results = []
    for L in args.L:
        try:
            r = run_one(L=L, dim=args.dim, qmax=args.qmax, lmax=args.lmax,
                        cutoff=args.cutoff, dis=args.dis, dis_type=args.dis_type,
                        torch_only=args.torch_only)
            all_results.append(r)
            out_path = _write_per_l_result_file(r, args)
            row("导出文件", str(out_path))
        except Exception as e:
            import traceback
            print(f"\n  [ERROR] L={L} 运行失败: {e}")
            traceback.print_exc()
            err_result = {
                "L": L,
                "n": L ** args.dim,
                "pass": False,
                "error": str(e),
                "t_torch": np.nan,
                "t_jax": np.nan,
                "torch_part1_init_s": np.nan,
                "torch_part2_ckpt_s": np.nan,
                "torch_forward_diag_s": np.nan,
                "torch_backward_liom_s": np.nan,
                "jax_part1_init_s": np.nan,
                "jax_part2_ckpt_s": np.nan,
                "jax_forward_diag_s": np.nan,
                "jax_backward_liom_s": np.nan,
            }
            all_results.append(err_result)
            out_path = _write_per_l_result_file(err_result, args)
            row("导出文件", str(out_path))

    # ── 汇总表 ────────────────────────────────────────────────────────────────
    print(f"\n{'═' * W}")
    print("  汇总")
    print(f"  {'─' * (W - 2)}")

    if args.torch_only:
        print(f"  {'L':>4}  {'n':>4}  {'off-diag':>10}  {'LIOM 迹':>9}  {'耗时(s)':>8}  状态")
        print(f"  {'─' * 52}")
        for r in all_results:
            if "error" in r:
                print(f"  {r['L']:>4}  {'':>4}  ERROR  {r.get('error','')[:38]}")
            else:
                status = "PASS" if r["pass"] else "WARN"
                print(f"  {r['L']:>4}  {r['n']:>4}  {r['offdiag_torch']:>10.2e}"
                      f"  {r['trace_torch']:>9.5f}  {r['t_torch']:>8.2f}  {status}")
    else:
        print(f"  {'L':>4}  {'n':>4}  {'eig_err':>9}  {'liom_err':>9}"
              f"  {'t_torch':>9}  {'t_jax':>9}  状态")
        print(f"  {'─' * 62}")
        for r in all_results:
            if "error" in r:
                print(f"  {r['L']:>4}  {'':>4}  ERROR  {r.get('error','')[:38]}")
            else:
                status = "PASS" if r["pass"] else "WARN/DIFF"
                print(f"  {r['L']:>4}  {r['n']:>4}  {r['eig_err']:>9.2e}"
                      f"  {r['liom_err']:>9.2e}"
                      f"  {r['t_torch']:>8.2f}s  {r['t_jax']:>8.2f}s  {status}")

    print(f"{'═' * W}")


if __name__ == "__main__":
    main()
