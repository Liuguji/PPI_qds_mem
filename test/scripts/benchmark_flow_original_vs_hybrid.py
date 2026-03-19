#!/usr/bin/env python3
"""
Benchmark flow_static_int vs flow_static_int_hybrid: run both with the same parameters,
collect phase timings (1.初始化, 2.迭代循环, 2.1-2.5, 2b.LIOM反向, 3.输出结果, 总时间),
and write results to Excel files (one per L value) under an output directory.

Output directory is created automatically if it does not exist.
Each file is named:  benchmark_L{L}_{timestamp}.xlsx

Usage:
  python test/scripts/benchmark_flow_original_vs_hybrid.py --L 4 6 8
  python test/scripts/benchmark_flow_original_vs_hybrid.py --L 4 --qmax 750 --lmax 75 --out-dir results/timing
"""

from __future__ import annotations

import os
import sys
import argparse
import time
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
CODE_DIR = REPO_ROOT / "code"
sys.path.insert(0, str(CODE_DIR))

# Phase keys for Excel (row order). 2a.LIOM正向 is skipped (not measured).
PHASE_ROW_ORDER = [
    "1.初始化",
    "2.迭代循环",
    "2.1_获取状态",
    "2.2_2.3_分离H2_Hint",
    "2.4_构建生成元",
    "2.5_矩阵变换",
    "2b.LIOM反向",
    "3.输出结果",
    "总时间",
]


def _build_params_and_hamiltonian(
    n: int,
    qmax: int,
    lmax: float,
    cutoff: float,
    method: str,
    norm: bool,
    force_steps: int | None,
    dis_type: str = "random",
    dis: list | float = None,
):
    """Build params dict and hamiltonian compatible with diag.CUT."""
    import numpy as np
    import jax.numpy as jnp
    import models.models as models

    dim = 2
    species = "spinless fermion"
    if dis is None:
        dis = [5.0]
    d = dis[0] if isinstance(dis, (list, tuple)) else dis
    delta = 0.1
    J = 1.0
    x = 0.0
    logflow = True

    params = {
        "n": n,
        "delta": delta,
        "J": J,
        "cutoff": cutoff,
        "dis": dis if isinstance(dis, (list, tuple)) else [dis],
        "dsymm": "spin",
        "NO_state": "SDW",
        "lmax": lmax,
        "qmax": qmax,
        "reps": 1,
        "norm": norm,
        "Hflow": True,
        "method": method,
        "intr": True,
        "dyn": False,
        "imbalance": True,
        "species": species,
        "LIOM": "bck",
        "dyn_MF": True,
        "logflow": logflow,
        "dis_type": dis_type,
        "x": x,
        "tlist": [],
        "store_flow": False,
        "ITC": False,
        "ladder": False,
        "order": 4,
        "dim": dim,
    }
    if force_steps is not None:
        params["_force_steps"] = force_steps

    ham = models.hamiltonian(species, dis_type, intr=True)
    ham.build(n, dim, d, J, x, delta=delta)

    num = jnp.zeros((n, n))
    num = num.at[n // 2, n // 2].set(1.0)
    num_int = jnp.zeros((n, n, n, n), dtype=jnp.float64)

    return params, ham, num, num_int


def _write_result(
    rows: list,
    out_path: Path,
    L: int,
    extra_info: dict | None = None,
):
    """Write timing rows to Excel (or CSV fallback). One file per L."""
    out_path = Path(out_path)
    write_csv = out_path.suffix.lower() == ".csv"

    if write_csv:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(f"# L={L}\n")
            if extra_info:
                for k, v in extra_info.items():
                    f.write(f"# {k}={v}\n")
            f.write("阶段,flow_static_int(s),flow_static_int_hybrid(s)\n")
            for r in rows:
                f.write(f"{r[0]},{r[1]},{r[2]}\n")
        print(f"  写入: {out_path}")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = f"L={L}"

        # Meta info rows
        if extra_info:
            for k, v in extra_info.items():
                ws.append([f"{k}: {v}"])
            ws.append([])

        # Header
        header = ["阶段", "flow_static_int (s)", "flow_static_int_hybrid (s)"]
        ws.append(header)
        hrow = ws.max_row
        for col in range(1, 4):
            cell = ws.cell(row=hrow, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for r in rows:
            ws.append([r[0], r[1], r[2]])

        # Column widths
        ws.column_dimensions[get_column_letter(1)].width = 28
        ws.column_dimensions[get_column_letter(2)].width = 24
        ws.column_dimensions[get_column_letter(3)].width = 30

        wb.save(str(out_path))
        print(f"  写入: {out_path}")

    except ImportError:
        csv_path = out_path.with_suffix(".csv")
        with open(csv_path, "w", encoding="utf-8") as f:
            f.write(f"# L={L}\n")
            if extra_info:
                for k, v in extra_info.items():
                    f.write(f"# {k}={v}\n")
            f.write("阶段,flow_static_int(s),flow_static_int_hybrid(s)\n")
            for r in rows:
                f.write(f"{r[0]},{r[1]},{r[2]}\n")
        print(f"  openpyxl 未安装，写入 CSV: {csv_path}")


def run_benchmark_one_L(
    L: int,
    dim: int = 2,
    qmax: int = 750,
    lmax: float = 75.0,
    cutoff: float | None = None,
    method: str = "einsum",
    norm: bool = False,
    force_steps: int | None = None,
    out_dir: Path | None = None,
    dis_type: str = "random",
    dis: list | float = None,
    timestamp: str | None = None,
) -> dict:
    """
    Run both flow functions for a single L value and write one result file.

    Returns dict with timing_orig, timing_hybrid, rows, out_path.
    """
    import core.diag as diag

    n = L ** dim
    if cutoff is None:
        cutoff = 1e-3

    print(f"\n{'='*60}")
    print(f"  L={L}  n={n}  qmax={qmax}  lmax={lmax}  method={method}")
    print(f"{'='*60}")

    params, ham, num, num_int = _build_params_and_hamiltonian(
        n=n, qmax=qmax, lmax=lmax, cutoff=cutoff, method=method,
        norm=norm, force_steps=force_steps, dis_type=dis_type, dis=dis,
    )

    # --- flow_static_int (standard) ---
    print(f"\n[L={L}] 运行 flow_static_int ...")
    os.environ["USE_CKPT"] = "0"
    t0 = time.perf_counter()
    flow_orig = diag.CUT(params, ham, num, num_int)
    t_orig_total = time.perf_counter() - t0
    timing_orig = flow_orig.pop("_timing", {}) or {}
    print(f"[L={L}] flow_static_int 完成，墙钟时间 {t_orig_total:.2f}s")

    # --- flow_static_int_hybrid ---
    print(f"\n[L={L}] 运行 flow_static_int_hybrid ...")
    os.environ["USE_CKPT"] = "hybrid"
    t0 = time.perf_counter()
    flow_hybrid = diag.CUT(params, ham, num, num_int)
    t_hybrid_total = time.perf_counter() - t0
    timing_hybrid = flow_hybrid.pop("_timing", {}) or {}
    print(f"[L={L}] flow_static_int_hybrid 完成，墙钟时间 {t_hybrid_total:.2f}s")

    # --- Build table rows ---
    rows = []
    for key in PHASE_ROW_ORDER:
        v_orig = timing_orig.get(key, "")
        v_hybrid = timing_hybrid.get(key, "")
        rows.append((key, v_orig, v_hybrid))

    # --- Output file ---
    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"benchmark_L{L}_{ts}.xlsx"
    if out_dir is not None:
        out_path = Path(out_dir) / filename
    else:
        out_path = REPO_ROOT / filename

    extra_info = {
        "L": L, "n": n, "dim": dim, "qmax": qmax, "lmax": lmax,
        "cutoff": cutoff, "method": method, "dis_type": dis_type,
        "timestamp": ts,
    }
    _write_result(rows, out_path, L=L, extra_info=extra_info)

    return {
        "L": L,
        "timing_orig": timing_orig,
        "timing_hybrid": timing_hybrid,
        "rows": rows,
        "out_path": str(out_path),
    }


def run_benchmark(
    L_list: list[int],
    dim: int = 2,
    qmax: int = 750,
    lmax: float = 75.0,
    cutoff: float | None = None,
    method: str = "einsum",
    norm: bool = False,
    force_steps: int | None = None,
    out_dir: Path | None = None,
    dis_type: str = "random",
    dis: list | float = None,
) -> list[dict]:
    """
    Run benchmarks for all L values in L_list.

    For each L:
      - Runs flow_static_int and flow_static_int_hybrid with identical parameters.
      - Writes one Excel file per L into out_dir (created if missing).

    All files in a single run share the same timestamp prefix so they sort together.
    """
    # Set environment before any core.diag import
    os.environ["BENCHMARK_FLOW_TIMING"] = "1"
    os.environ["USE_JIT_FLOW"] = "0"          # use scipy solver so int_ode is a Python callback
    os.environ["PYFLOW_ADAPTIVE_GRID"] = "0"  # disable adaptive grid for reproducibility
    if force_steps is not None:
        os.environ["PYFLOW_FORCE_STEPS"] = str(force_steps)

    # Resolve output directory and create it if needed
    if out_dir is None:
        out_dir = REPO_ROOT / "benchmark_results"
    out_dir = Path(out_dir)
    if not out_dir.exists():
        out_dir.mkdir(parents=True, exist_ok=True)
        print(f"创建输出目录: {out_dir}")
    else:
        print(f"输出目录: {out_dir}")

    # Shared timestamp so all files from this run sort together
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    results = []
    for L in L_list:
        result = run_benchmark_one_L(
            L=L, dim=dim, qmax=qmax, lmax=lmax, cutoff=cutoff,
            method=method, norm=norm, force_steps=force_steps,
            out_dir=out_dir, dis_type=dis_type, dis=dis,
            timestamp=ts,
        )
        results.append(result)

    print(f"\n{'='*60}")
    print(f"全部完成，共 {len(L_list)} 个 L 值，结果保存在: {out_dir}")
    print(f"{'='*60}")
    return results


def main():
    parser = argparse.ArgumentParser(
        description="Benchmark flow_static_int vs flow_static_int_hybrid for one or more L values.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python benchmark_flow_original_vs_hybrid.py --L 4
  python benchmark_flow_original_vs_hybrid.py --L 4 6 8
  python benchmark_flow_original_vs_hybrid.py --L 4 6 --qmax 500 --out-dir results/timing
        """,
    )
    parser.add_argument(
        "--L", type=int, nargs="+", default=[4],
        help="线性系统大小列表（n = L**dim），可同时指定多个，例如 --L 4 6 8",
    )
    parser.add_argument("--dim", type=int, default=2, help="空间维度（默认 2）")
    parser.add_argument("--qmax", type=int, default=750, help="最大流方程步数（默认 750）")
    parser.add_argument("--lmax", type=float, default=75.0, help="最大流时间（默认 75）")
    parser.add_argument("--cutoff", type=float, default=None, help="非对角截断值（默认 1e-3）")
    parser.add_argument(
        "--method", type=str, default="einsum",
        help="张量收缩方法：einsum / tensordot / jit / vec（默认 einsum）",
    )
    parser.add_argument("--norm", action="store_true", help="启用 normal ordering")
    parser.add_argument(
        "--force-steps", type=int, default=None, dest="force_steps",
        help="强制固定步数（忽略 cutoff），用于公平比较",
    )
    parser.add_argument(
        "--dis-type", type=str, default="random", dest="dis_type",
        help="无序类型（默认 random）",
    )
    parser.add_argument(
        "--out-dir", type=str, default=None, dest="out_dir",
        help="结果输出目录（不存在时自动创建，默认 <repo>/benchmark_results）",
    )

    args = parser.parse_args()

    out_dir = Path(args.out_dir) if args.out_dir else None

    run_benchmark(
        L_list=args.L,
        dim=args.dim,
        qmax=args.qmax,
        lmax=args.lmax,
        cutoff=args.cutoff,
        method=args.method,
        norm=args.norm,
        force_steps=args.force_steps,
        out_dir=out_dir,
        dis_type=args.dis_type,
    )


if __name__ == "__main__":
    main()
